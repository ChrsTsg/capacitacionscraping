#scraping
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options
#excel
import openpyxl
#espera por segundo
import time
#correo
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Configuración de Selenium
firefox_options = Options()
#firefox_options.add_argument("--headless") # Modo headless significa que el navegador no abrirá una ventana de GUI
service = Service(GeckoDriverManager().install())
driver = webdriver.Firefox(service=service, options=firefox_options)
wait = WebDriverWait(driver, 10) # Tiempo de espera para que Selenium espere a que los elementos estén disponibles

# Configuración inicial de Excel
try:
    wb = openpyxl.load_workbook('ofertas.xlsx') # Intenta abrir un archivo existente
    sheet = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook() # Si el archivo no existe, crea uno nuevo
    sheet = wb.active
    sheet.title = 'ofertas'
    sheet.append(['Nombre', 'precio', 'descuento']) # Agrega encabezados a la hoja de Excel

url = 'https://www.mercadolibre.cl/ofertas#nav-header'
driver.get(url) # Abre la URL en el navegador
time.sleep(5) # Espera para asegurar que la página cargue completamente

# Define los XPaths de los elementos a capturar
xpath1 = '/html/body/main/div[2]/div[3]/div[2]/div[1]/div[1]/div/div[1]/div[2]/div[1]/div/div[2]/h1'
xpath2 = '/html/body/main/div[2]/div[2]/div/ol/li[{}]/div/a/div/div[2]/div[2]/div/div/span/span[2]'
xpath3 = '/html/body/main/div[2]/div[2]/div/ol/li[{}]/div/a/div/div[2]/div[2]/span'
xpath4 = '/html/body/main/div[2]/div[2]/div/ol/li[{}]/div/a/div/div[1]/img'

# Función para realizar el web scraping y guardar los datos en Excel
def buscar_en_mercado_y_guardar_datos():
    max_intentos = 100  # Número máximo de intentos para encontrar un elemento
    intentos = 0
    i = 1 
    while True:
        if i >= 10:
            print('termino')
            break
        print(i)
        
        try:
            print('click')
            click1 = wait.until(EC.element_to_be_clickable((By.XPATH, xpath4.format(i))))
            click1.click()
            time.sleep(2)
        except:
            clickcookie = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/div/div[2]/button[1]')))
            if clickcookie:
                clickcookie.click()

        try:
            
            fila = sheet.max_row + 1

            texto1 = wait.until(EC.visibility_of_element_located((By.XPATH, xpath1))).text
            print(texto1)
            print('volviendo')
            driver.back()
            time.sleep(2)
            texto2 = wait.until(EC.visibility_of_element_located((By.XPATH, xpath2.format(i)))).text
            print(texto2)
            texto3 = wait.until(EC.visibility_of_element_located((By.XPATH, xpath3.format(i)))).text
            print(texto3)
            # Guarda los datos en Excel
            sheet.cell(row=fila, column=1, value=texto1)
            sheet.cell(row=fila, column=2, value=texto2)
            sheet.cell(row=fila, column=3, value=texto3)
            i += 1
            intentos = 0  # Restablece intentos después de un éxito
            print('guardando excel')
            wb.save('ofertas.xlsx')
        except Exception as e:
            print(e)
            driver.back()
            intentos += 1
            i += 1
            if intentos >= max_intentos:
                print(f"No se pudo encontrar el elemento después de {max_intentos} intentos. Error: {e}")
                wb.save('ofertas.xlsx')
                break  # Sale del bucle si se alcanza el número máximo de intentos

# Ejecuta la función de scraping y guarda el archivo Excel
try:
    buscar_en_mercado_y_guardar_datos()
except Exception as e:
    time.sleep(5)
    print(f'Error general: {e}')
    driver.quit()
finally:
    driver.quit()
    wb.save('ofertas.xlsx')

# Función para enviar un correo electrónico con el archivo adjunto
def send_email_with_attachment(from_addr, to_addr, cc_addr, subject, password, file_path):
    msg = MIMEMultipart()
    msg['From'] = from_addr
    msg['To'] = ', '.join(to_addr)
    msg['CC'] = ', '.join(cc_addr)
    msg['Subject'] = subject

    # Adjunta el archivo
    attachment = open(file_path, "rb")
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= " + file_path)
    msg.attach(part)

    # Crea una conexión segura con el servidor SMTP utilizando SSL
    server = smtplib.SMTP('smtp.office365.com', 587)  # Ajusta esto a tu servidor SMTP si es necesario
    server.starttls()
    # Inicia sesión en la cuenta de correo
    server.login(from_addr, password)

    # Envía el correo
    server.send_message(msg)

    # Cierra la conexión con el servidor
    server.quit()

# Ejemplo de uso
from_addr = "ciat.0313@gmail.com"
to_addr = ["chris.alc.13@gmail.com"]
cc_addr = [""]
subject = "Asunto del Email"
password = "Chris.572"
file_path = "ofertas.xlsx"
try:
    send_email_with_attachment(from_addr, to_addr, cc_addr, subject, password, file_path)
except Exception as e:
    print(f'error correo {e}')